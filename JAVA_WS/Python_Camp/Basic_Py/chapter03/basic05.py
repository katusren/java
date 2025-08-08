import openpyxl
'''
# 부분 로딩
wb = openpyxl.load_workbook('sample.xlsx')
sheet = wb['Sheet1']

print(sheet.max_column,sheet.max_row) # 총 컬럼 총 행의 수
print(sheet.cell(row=1,column=1).value)
print(sheet.cell(row=2,column=1).value)

wb.close()

print()
'''

# 행 단위로 로딩
wb = openpyxl.load_workbook('sample.xlsx')
sheet = wb['Sheet1']
cells=sheet['A2':'C4']
for row in cells:
    for cells in row:
        print(cells.value)
    print('-'*10)
wb.close()